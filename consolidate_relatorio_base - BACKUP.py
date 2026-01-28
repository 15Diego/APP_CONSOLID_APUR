
from __future__ import annotations

import os
import re
import threading
import logging
from dataclasses import dataclass
from datetime import datetime
from typing import Iterable, List, Optional, Tuple

import pandas as pd

import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_SHEET_NAME = "Relatorio Base"


# -----------------------------
# Utilitários de normalização
# -----------------------------

def _canon_text(value: str) -> str:
    """Normaliza texto para comparação (caixa baixa, espaços colapsados)."""
    s = "" if value is None else str(value)
    s = re.sub(r"\s+", " ", s.strip())
    return s.lower()


def _normalize_column_name(col: object) -> str:
    """Normaliza nome de coluna (string limpa)."""
    if col is None or (isinstance(col, float) and pd.isna(col)):
        return ""
    s = str(col)
    s = re.sub(r"\s+", " ", s.strip())
    return s


def _make_unique(names: List[str]) -> List[str]:
    """Garante unicidade de nomes de coluna, aplicando sufixos _2, _3, ..."""
    seen: dict[str, int] = {}
    out: List[str] = []
    for n in names:
        base = n
        if base not in seen:
            seen[base] = 1
            out.append(base)
            continue
        seen[base] += 1
        out.append(f"{base}_{seen[base]}")
    return out


def _is_filled_cell(v: object) -> bool:
    if v is None:
        return False
    if isinstance(v, float) and pd.isna(v):
        return False
    s = str(v).strip()
    if s == "":
        return False
    if s.lower() in {"nan", "none", "<na>"}:
        return False
    return True


def _looks_numeric(s: str) -> bool:
    # números com separadores comuns
    s = s.strip()
    return bool(re.fullmatch(r"[-+]?\d{1,3}([\.,]\d{3})*([\.,]\d+)?", s) or re.fullmatch(r"[-+]?\d+([\.,]\d+)?", s))


# -----------------------------
# Leitura robusta do Excel
# -----------------------------

def resolve_sheet_name(file_path: str, preferred_sheet: str) -> str:
    """Resolve o nome real da aba no arquivo, de forma tolerante."""
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    preferred_canon = _canon_text(preferred_sheet)

    # match exato primeiro
    if preferred_sheet in xls.sheet_names:
        return preferred_sheet

    # match tolerante
    for s in xls.sheet_names:
        if _canon_text(s) == preferred_canon:
            return s

    raise ValueError(
        f"Aba '{preferred_sheet}' não encontrada em '{os.path.basename(file_path)}'. "
        f"Abas disponíveis: {', '.join(xls.sheet_names)}"
    )


def detect_header_row(
    file_path: str,
    sheet_name: str,
    max_rows: int = 30,
) -> int:
    """Detecta a linha (0-based) do cabeçalho analisando as primeiras linhas."""

    # lê prévia sem cabeçalho
    try:
        preview = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            nrows=max_rows,
            engine="openpyxl",
            dtype="string",
        )
    except TypeError:
        # fallback para versões mais antigas
        preview = pd.read_excel(
            file_path,
            sheet_name=sheet_name,
            header=None,
            nrows=max_rows,
            engine="openpyxl",
            dtype=str,
        )

    best_row = 0
    best_score = -1.0

    for i in range(len(preview)):
        row = preview.iloc[i].tolist()
        filled = [v for v in row if _is_filled_cell(v)]
        non_empty = len(filled)
        if non_empty < 2:
            continue

        # unicidade e penalização de linha "muito numérica" (tende a ser dados)
        filled_str = [str(v).strip() for v in filled]
        unique_count = len(set(filled_str))
        numeric_ratio = 0.0
        if filled_str:
            numeric_ratio = sum(_looks_numeric(x) for x in filled_str) / len(filled_str)

        # score: mais preenchida + mais única; penaliza se muito numérica
        score = non_empty + 0.5 * unique_count - 2.0 * numeric_ratio

        if score > best_score:
            best_score = score
            best_row = i

    return best_row


@dataclass
class ReadResult:
    df: Optional[pd.DataFrame]
    arquivo: str
    aba: Optional[str]
    header_row_0based: Optional[int]
    linhas: int
    colunas: int
    status: str
    erro: Optional[str] = None


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Remove colunas vazias/Unnamed, normaliza nomes e remove linhas totalmente vazias."""

    # normaliza nomes
    cols = [_normalize_column_name(c) for c in df.columns]

    # remove colunas sem nome ou Unnamed
    keep_mask = []
    cleaned_cols = []
    for c in cols:
        if c == "":
            keep_mask.append(False)
            continue
        if _canon_text(c).startswith("unnamed"):
            keep_mask.append(False)
            continue
        keep_mask.append(True)
        cleaned_cols.append(c)

    df = df.loc[:, keep_mask]

    # unicidade de nomes
    cleaned_cols = _make_unique(cleaned_cols)
    df.columns = cleaned_cols

    # remove linhas totalmente vazias
    df = df.dropna(axis=0, how="all")

    # remove colunas totalmente vazias
    df = df.dropna(axis=1, how="all")

    return df


def ler_planilha_robusta(
    file_path: str,
    preferred_sheet: str,
    auto_detect_header: bool = True,
    header_row_0based: Optional[int] = None,
    read_as_text: bool = True,
    adicionar_auditoria: bool = True,
) -> ReadResult:
    """Lê uma planilha (um arquivo) e retorna DataFrame + metadados, sem abortar o processo inteiro."""

    arquivo = os.path.basename(file_path)

    try:
        aba_real = resolve_sheet_name(file_path, preferred_sheet)

        if header_row_0based is None:
            header = detect_header_row(file_path, aba_real) if auto_detect_header else 1
        else:
            header = header_row_0based

        dtype_arg = "string" if read_as_text else None
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=aba_real,
                header=header,
                engine="openpyxl",
                dtype=dtype_arg,
            )
        except TypeError:
            # pandas sem suporte a dtype="string" em read_excel
            df = pd.read_excel(
                file_path,
                sheet_name=aba_real,
                header=header,
                engine="openpyxl",
                dtype=str if read_as_text else None,
            )

        df = _clean_dataframe(df)

        # auditoria / rastreabilidade
        if adicionar_auditoria:
            # linha 1-based no Excel: cabeçalho está em header+1, dados começam em header+2
            linha_ini = header + 2
            df.insert(0, "ARQUIVO_ORIGEM", arquivo)
            df.insert(1, "ABA_ORIGEM", aba_real)
            df.insert(2, "HEADER_LINHA", header + 1)  # 1-based
            df.insert(3, "LINHA_ORIGEM_EXCEL", pd.Series(range(linha_ini, linha_ini + len(df)), dtype="Int64"))

        return ReadResult(
            df=df,
            arquivo=arquivo,
            aba=aba_real,
            header_row_0based=header,
            linhas=int(len(df)),
            colunas=int(df.shape[1]),
            status="OK",
        )

    except Exception as e:
        return ReadResult(
            df=None,
            arquivo=arquivo,
            aba=None,
            header_row_0based=None,
            linhas=0,
            colunas=0,
            status="FALHA",
            erro=str(e),
        )


# -----------------------------
# Consolidação
# -----------------------------

def _colunas_final(planilhas: List[pd.DataFrame]) -> List[str]:
    """Define colunas finais preservando a ordem do primeiro arquivo e adicionando extras por 1ª aparição."""
    if not planilhas:
        return []

    final: List[str] = []
    seen: set[str] = set()

    def add(c: str) -> None:
        if c not in seen:
            seen.add(c)
            final.append(c)

    for c in planilhas[0].columns:
        add(c)

    for df in planilhas[1:]:
        for c in df.columns:
            add(c)

    return final


def consolidar_planilhas(planilhas: List[pd.DataFrame]) -> pd.DataFrame:
    """Consolida uma lista de DataFrames unificando colunas."""
    if not planilhas:
        return pd.DataFrame()

    cols = _colunas_final(planilhas)
    aligned = [df.reindex(columns=cols) for df in planilhas]
    return pd.concat(aligned, ignore_index=True)


# -----------------------------
# Saída Excel (com formatação)
# -----------------------------

def _aplicar_formatacao_worksheet(ws, table_name: str) -> None:
    """Aplica congelação, filtro, larguras e formata como tabela."""

    if ws.max_row < 1 or ws.max_column < 1:
        return

    # congelar cabeçalho
    ws.freeze_panes = "A2"

    # filtro
    ws.auto_filter.ref = ws.dimensions

    # ajustar larguras (amostra: cabeçalho + primeiras 200 linhas)
    max_row_sample = min(ws.max_row, 201)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, max_row_sample + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(max_len + 2, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, width)

    # tabela (se houver ao menos cabeçalho + 1 linha)
    if ws.max_row >= 2:
        last_col = get_column_letter(ws.max_column)
        ref = f"A1:{last_col}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style

        # evita duplicidade se re-salvar
        if not any(t.displayName == table_name for t in ws._tables):
            ws.add_table(table)


def salvar_excel(
    df_dados: pd.DataFrame,
    df_resumo: pd.DataFrame,
    output_path: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
    formatar: bool = True,
) -> None:
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_dados.to_excel(writer, sheet_name=sheet_name, index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)

        if formatar:
            wb = writer.book
            ws_dados = wb[sheet_name]
            ws_resumo = wb["Resumo"]
            _aplicar_formatacao_worksheet(ws_dados, table_name="DadosConsolidados")
            _aplicar_formatacao_worksheet(ws_resumo, table_name="ResumoConsolidacao")


# -----------------------------
# Função principal reutilizável
# -----------------------------

def consolidar(
    arquivos: Iterable[str],
    output_path: str,
    preferred_sheet: str = DEFAULT_SHEET_NAME,
    auto_detect_header: bool = True,
    header_row_0based: Optional[int] = None,
    read_as_text: bool = True,
    adicionar_auditoria: bool = True,
    formatar_saida: bool = True,
    logger: Optional[logging.Logger] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Consolida múltiplos arquivos Excel e salva o resultado.

    Retorna (df_consolidado, df_resumo).
    """

    log = logger or logging.getLogger("consolidacao")

    resultados: List[ReadResult] = []
    dfs_ok: List[pd.DataFrame] = []

    for p in arquivos:
        r = ler_planilha_robusta(
            file_path=p,
            preferred_sheet=preferred_sheet,
            auto_detect_header=auto_detect_header,
            header_row_0based=header_row_0based,
            read_as_text=read_as_text,
            adicionar_auditoria=adicionar_auditoria,
        )
        resultados.append(r)
        if r.status == "OK" and r.df is not None:
            dfs_ok.append(r.df)
            log.info("OK: %s (%s linhas)", r.arquivo, r.linhas)
        else:
            log.error("FALHA: %s -> %s", r.arquivo, r.erro)

    df_consolidado = consolidar_planilhas(dfs_ok)

    df_resumo = pd.DataFrame(
        [
            {
                "arquivo": r.arquivo,
                "status": r.status,
                "aba": r.aba,
                "header_linha": (r.header_row_0based + 1) if r.header_row_0based is not None else None,
                "linhas": r.linhas,
                "colunas": r.colunas,
                "erro": r.erro,
            }
            for r in resultados
        ]
    )

    if df_consolidado.empty and not any(r.status == "OK" for r in resultados):
        raise RuntimeError(
            "Nenhum arquivo foi consolidado com sucesso. Verifique a aba/cabeçalho e o resumo de erros."
        )

    salvar_excel(
        df_dados=df_consolidado,
        df_resumo=df_resumo,
        output_path=output_path,
        sheet_name=preferred_sheet,
        formatar=formatar_saida,
    )

    return df_consolidado, df_resumo


# -----------------------------
# Interface gráfica (Tkinter)
# -----------------------------

class App(tk.Tk):
    def __init__(self) -> None:
        super().__init__()

        self.title("Consolidar Relatório Base")
        self.geometry("780x520")
        self.minsize(780, 520)

        self.selected_files: List[str] = []

        # Opções
        self.var_sheet = tk.StringVar(value=DEFAULT_SHEET_NAME)
        self.var_auto_header = tk.BooleanVar(value=True)
        self.var_header_manual = tk.IntVar(value=2)  # 1-based para o usuário
        self.var_read_text = tk.BooleanVar(value=True)
        self.var_auditoria = tk.BooleanVar(value=True)
        self.var_formatar = tk.BooleanVar(value=True)

        self._build_ui()

    def _build_ui(self) -> None:
        frm_top = ttk.Frame(self, padding=10)
        frm_top.pack(fill="x")

        # Linha: aba
        ttk.Label(frm_top, text="Nome da aba:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm_top, textvariable=self.var_sheet, width=30).grid(row=0, column=1, sticky="w", padx=6)

        # Linha: header
        ttk.Checkbutton(
            frm_top,
            text="Detectar cabeçalho automaticamente",
            variable=self.var_auto_header,
            command=self._toggle_header_manual,
        ).grid(row=0, column=2, sticky="w", padx=10)

        ttk.Label(frm_top, text="Cabeçalho (linha):").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.spn_header = ttk.Spinbox(frm_top, from_=1, to=200, textvariable=self.var_header_manual, width=6)
        self.spn_header.grid(row=1, column=1, sticky="w", padx=6, pady=(8, 0))

        # Opções
        frm_opts = ttk.Frame(self, padding=(10, 0, 10, 10))
        frm_opts.pack(fill="x")

        ttk.Checkbutton(frm_opts, text="Ler tudo como texto (preserva zeros à esquerda)", variable=self.var_read_text).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Checkbutton(frm_opts, text="Adicionar colunas de auditoria (origem/linha)", variable=self.var_auditoria).grid(
            row=1, column=0, sticky="w", pady=(4, 0)
        )
        ttk.Checkbutton(frm_opts, text="Formatar saída (filtro, congelar, larguras, tabela)", variable=self.var_formatar).grid(
            row=2, column=0, sticky="w", pady=(4, 0)
        )

        # Botões
        frm_btn = ttk.Frame(self, padding=10)
        frm_btn.pack(fill="x")

        self.btn_select = ttk.Button(frm_btn, text="Selecionar arquivos (.xlsx)", command=self._select_files)
        self.btn_select.pack(side="left")

        self.btn_run = ttk.Button(frm_btn, text="Consolidar e salvar", command=self._run)
        self.btn_run.pack(side="left", padx=8)

        # Lista de arquivos
        frm_list = ttk.Frame(self, padding=(10, 0, 10, 10))
        frm_list.pack(fill="both", expand=True)

        ttk.Label(frm_list, text="Arquivos selecionados:").pack(anchor="w")

        self.listbox = tk.Listbox(frm_list, height=10)
        self.listbox.pack(fill="both", expand=True, pady=(4, 0))

        # Barra de progresso
        frm_status = ttk.Frame(self, padding=10)
        frm_status.pack(fill="x")

        self.progress = ttk.Progressbar(frm_status, mode="determinate")
        self.progress.pack(fill="x")

        self.lbl_status = ttk.Label(frm_status, text="")
        self.lbl_status.pack(anchor="w", pady=(6, 0))

        self._toggle_header_manual()

    def _toggle_header_manual(self) -> None:
        # Se auto-detect está ligado, desabilita spinbox
        state = "disabled" if self.var_auto_header.get() else "normal"
        self.spn_header.configure(state=state)

    def _select_files(self) -> None:
        paths = filedialog.askopenfilenames(
            title="Selecione as planilhas mensais a consolidar",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if not paths:
            return

        self.selected_files = list(paths)
        self.listbox.delete(0, tk.END)
        for p in self.selected_files:
            self.listbox.insert(tk.END, p)

        self.lbl_status.configure(text=f"{len(self.selected_files)} arquivo(s) selecionado(s).")

    def _run(self) -> None:
        if not self.selected_files:
            messagebox.showwarning("Atenção", "Selecione ao menos um arquivo .xlsx.")
            return

        output_path = filedialog.asksaveasfilename(
            title="Salvar arquivo consolidado como",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if not output_path:
            return

        # logger para arquivo
        log_path = os.path.splitext(output_path)[0] + ".log"
        logger = logging.getLogger("consolidacao")
        logger.setLevel(logging.INFO)
        logger.handlers.clear()
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        logger.addHandler(fh)

        # configura UI
        self.btn_select.configure(state="disabled")
        self.btn_run.configure(state="disabled")
        self.progress.configure(maximum=len(self.selected_files), value=0)
        self.lbl_status.configure(text="Iniciando...")

        args = {
            "arquivos": self.selected_files,
            "output_path": output_path,
            "preferred_sheet": self.var_sheet.get().strip() or DEFAULT_SHEET_NAME,
            "auto_detect_header": self.var_auto_header.get(),
            "header_row_0based": None if self.var_auto_header.get() else max(0, int(self.var_header_manual.get()) - 1),
            "read_as_text": self.var_read_text.get(),
            "adicionar_auditoria": self.var_auditoria.get(),
            "formatar_saida": self.var_formatar.get(),
            "logger": logger,
        }

        def worker() -> None:
            try:
                # progresso por arquivo: fazemos leitura manual aqui para atualizar barra
                results: List[ReadResult] = []
                dfs_ok: List[pd.DataFrame] = []

                for idx, p in enumerate(self.selected_files, start=1):
                    r = ler_planilha_robusta(
                        file_path=p,
                        preferred_sheet=args["preferred_sheet"],
                        auto_detect_header=args["auto_detect_header"],
                        header_row_0based=args["header_row_0based"],
                        read_as_text=args["read_as_text"],
                        adicionar_auditoria=args["adicionar_auditoria"],
                    )
                    results.append(r)
                    if r.status == "OK" and r.df is not None:
                        dfs_ok.append(r.df)
                    args["logger"].info("%s: %s", r.status, r.arquivo)

                    # update UI
                    self.after(0, self._set_progress, idx, f"Processado: {r.arquivo} ({r.status})")

                df_consolidado = consolidar_planilhas(dfs_ok)
                df_resumo = pd.DataFrame(
                    [
                        {
                            "arquivo": r.arquivo,
                            "status": r.status,
                            "aba": r.aba,
                            "header_linha": (r.header_row_0based + 1) if r.header_row_0based is not None else None,
                            "linhas": r.linhas,
                            "colunas": r.colunas,
                            "erro": r.erro,
                        }
                        for r in results
                    ]
                )

                if df_consolidado.empty and not any(r.status == "OK" for r in results):
                    raise RuntimeError("Nenhum arquivo foi consolidado com sucesso. Verifique o LOG e a aba 'Resumo'.")

                salvar_excel(
                    df_dados=df_consolidado,
                    df_resumo=df_resumo,
                    output_path=args["output_path"],
                    sheet_name=args["preferred_sheet"],
                    formatar=args["formatar_saida"],
                )

                ok_count = sum(r.status == "OK" for r in results)
                fail_count = sum(r.status != "OK" for r in results)
                total_rows = int(df_consolidado.shape[0])

                msg = (
                    f"Consolidação concluída.\n\n"
                    f"Arquivos OK: {ok_count}\n"
                    f"Arquivos com falha: {fail_count}\n"
                    f"Linhas consolidadas: {total_rows}\n\n"
                    f"Saída: {args['output_path']}\n"
                    f"Log: {log_path}"
                )
                self.after(0, self._finish_ok, msg)

            except Exception as e:
                self.after(0, self._finish_error, str(e), log_path)

        threading.Thread(target=worker, daemon=True).start()

    def _set_progress(self, value: int, status: str) -> None:
        self.progress.configure(value=value)
        self.lbl_status.configure(text=status)

    def _finish_ok(self, msg: str) -> None:
        self.btn_select.configure(state="normal")
        self.btn_run.configure(state="normal")
        self.lbl_status.configure(text="Concluído.")
        messagebox.showinfo("Consolidação concluída", msg)

    def _finish_error(self, err: str, log_path: str) -> None:
        self.btn_select.configure(state="normal")
        self.btn_run.configure(state="normal")
        self.lbl_status.configure(text="Erro.")
        messagebox.showerror(
            "Erro na consolidação",
            f"Ocorreu um erro:\n{err}\n\nVerifique o log em:\n{log_path}",
        )


def main() -> None:
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
