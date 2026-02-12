
from __future__ import annotations

import os
import re
import threading
import logging
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from enum import Enum
from typing import Dict, Iterable, List, Optional, Tuple, Any

import pandas as pd

try:
    import tkinter as tk
    from tkinter import filedialog, messagebox, ttk
except ImportError:
    tk = None
    filedialog = None
    messagebox = None
    ttk = None

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet


# ========================================
# Constantes de Configuração
# ========================================

# Nome padrão da aba
DEFAULT_SHEET_NAME = "Relatorio Base"

# Configurações de detecção de cabeçalho
HEADER_DETECTION_ROWS = 30
MIN_FILLED_CELLS_FOR_HEADER = 2
NUMERIC_PENALTY_WEIGHT = 2.0
UNIQUENESS_BONUS = 0.5

# Configurações de formatação Excel
MAX_COLUMN_WIDTH = 60
MIN_COLUMN_WIDTH = 10
COLUMN_PADDING = 2
WIDTH_SAMPLE_ROWS = 200
MAX_COLUMN_NAME_LENGTH = 255


# ========================================
# Enumerações e Classes de Configuração
# ========================================

class AuditColumn(str, Enum):
    """Colunas de auditoria adicionadas aos dados consolidados."""
    SOURCE_FILE = "ARQUIVO_ORIGEM"
    SOURCE_SHEET = "ABA_ORIGEM"
    HEADER_LINE = "HEADER_LINHA"
    ORIGINAL_ROW = "LINHA_ORIGEM_EXCEL"


@dataclass
class ConsolidationConfig:
    """Configurações centralizadas para consolidação."""
    preferred_sheet: str = DEFAULT_SHEET_NAME
    auto_detect_header: bool = True
    header_row_0based: Optional[int] = None
    read_as_text: bool = True
    add_audit_columns: bool = True
    format_output: bool = True
    max_detection_rows: int = HEADER_DETECTION_ROWS


# -----------------------------
# Utilitários de normalização
# -----------------------------

def _canon_text(value: str) -> str:
    """
    Normaliza texto para comparação insensível a caso e espaços.
    
    Args:
        value: Texto a ser normalizado. Pode ser qualquer tipo que seja
               conversível para string.
    
    Returns:
        String normalizada em lowercase com espaços colapsados em único espaço.
        Retorna string vazia se value for None.
    
    Examples:
        >>> _canon_text("  HELLO   WORLD  ")
        'hello world'
        >>> _canon_text(None)
        ''
    """
    s = "" if value is None else str(value)
    s = re.sub(r"\s+", " ", s.strip())
    return s.lower()


def _normalize_column_name(col: object) -> str:
    """
    Normaliza nome de coluna removendo espaços extras e limitando tamanho.
    
    Args:
        col: Nome da coluna a ser normalizado. Pode ser string, número ou None.
    
    Returns:
        String normalizada com espaços colapsados, limitada ao tamanho máximo
        permitido pelo Excel. Retorna string vazia se col for None ou NaN.
    
    Examples:
        >>> _normalize_column_name("  Nome  Completo  ")
        'Nome Completo'
        >>> _normalize_column_name(None)
        ''
    """
    if col is None or (isinstance(col, float) and pd.isna(col)):
        return ""
    s = str(col).strip()
    s = re.sub(r"\s+", " ", s)
    # Limita tamanho ao máximo permitido pelo Excel
    return s[:MAX_COLUMN_NAME_LENGTH]


def _make_unique(names: List[str]) -> List[str]:
    """
    Garante unicidade de nomes de coluna aplicando sufixos _2, _3, etc.
    
    Args:
        names: Lista de nomes de colunas potencialmente duplicados.
    
    Returns:
        Lista de nomes únicos. Duplicatas recebem sufixos _2, _3, ...
    
    Examples:
        >>> _make_unique(["col", "col", "col"])
        ['col', 'col_2', 'col_3']
        >>> _make_unique(["a", "b", "a"])
        ['a', 'b', 'a_2']
    """
    seen: Dict[str, int] = {}
    out: List[str] = []
    for n in names:
        base = n
        if base not in seen:
            seen[base] = 1
            out.append(base)
            continue
        seen[base] += 1
        out.append(f"{base}_{seen[base]}")
    return out


def _is_filled_cell(v: object) -> bool:
    """
    Verifica se uma célula contém valor preenchido (não vazio ou nulo).
    
    Args:
        v: Valor da célula a ser verificado.
    
    Returns:
        True se a célula contém valor útil, False caso contrário.
        Considera None, NaN, strings vazias e literais como vazios.
    
    Examples:
        >>> _is_filled_cell("texto")
        True
        >>> _is_filled_cell(None)
        False
        >>> _is_filled_cell("")
        False
    """
    if v is None:
        return False
    if isinstance(v, float) and pd.isna(v):
        return False
    s = str(v).strip()
    if s == "":
        return False
    if s.lower() in {"nan", "none", "<na>"}:
        return False
    return True


def _looks_numeric(s: str) -> bool:
    """
    Verifica se uma string parece representar um número.
    
    Args:
        s: String a ser verificada.
    
    Returns:
        True se a string parece numérica (inteiro ou decimal com
        separadores comuns), False caso contrário.
    
    Examples:
        >>> _looks_numeric("123.45")
        True
        >>> _looks_numeric("1.234,56")
        True
        >>> _looks_numeric("abc")
        False
    """
    if not s or not isinstance(s, str):
        return False
    s = s.strip()
    return bool(
        re.fullmatch(r"[-+]?\d{1,3}([.,]\d{3})*([.,]\d+)?", s) or 
        re.fullmatch(r"[-+]?\d+([.,]\d+)?", s)
    )


# -----------------------------
# Leitura robusta do Excel
# -----------------------------

def resolve_sheet_name(file_path: str, preferred_sheet: str) -> str:
    """
    Resolve o nome real da aba no arquivo Excel de forma tolerante.
    
    Args:
        file_path: Caminho completo do arquivo Excel.
        preferred_sheet: Nome preferencial da aba (pode ter variações de case/espaços).
    
    Returns:
        Nome real da aba encontrada no arquivo.
    
    Raises:
        ValueError: Se a aba não for encontrada, com lista de abas disponíveis.
    
    Examples:
        >>> resolve_sheet_name("data.xlsx", "relatorio base")
        'Relatorio Base'
    """
    xls = pd.ExcelFile(file_path, engine="openpyxl")
    preferred_canon = _canon_text(preferred_sheet)

    # match exato primeiro
    if preferred_sheet in xls.sheet_names:
        return preferred_sheet

    # match tolerante
    for s in xls.sheet_names:
        if _canon_text(s) == preferred_canon:
            return s

    # Mensagem de erro formatada e informativa
    raise ValueError(
        f"Aba '{preferred_sheet}' não encontrada.\n"
        f"Arquivo: {os.path.basename(file_path)}\n"
        f"Abas disponíveis:\n" + 
        "\n".join(f"  • {s}" for s in xls.sheet_names)
    )


def detect_header_row(
    file_path: str,
    sheet_name: str,
    max_rows: int = HEADER_DETECTION_ROWS,
) -> int:
    """
    Detecta a linha (0-based) do cabeçalho usando openpyxl direto.
    
    Esta versão é mais eficiente que carregar um DataFrame completo,
    pois lê apenas as linhas necessárias para detecção.
    
    O algoritmo analisa as primeiras linhas buscando por:
    - Maior número de células preenchidas
    - Maior unicidade de valores (colunas com nomes únicos)
    - Penaliza linhas muito numéricas (provavelmente dados, não cabeçalho)
    
    Args:
        file_path: Caminho completo do arquivo Excel.
        sheet_name: Nome da aba a analisar.
        max_rows: Número máximo de linhas a analisar (padrão: HEADER_DETECTION_ROWS).
    
    Returns:
        Índice (0-based) da linha que provavelmente contém o cabeçalho.
    
    Examples:
        >>> detect_header_row("data.xlsx", "Sheet1")
        1  # Segunda linha do Excel (1-based = 2)
    """
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]
    
    best_row = 0
    best_score = -1.0
    
    for i, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
        filled = [v for v in row if _is_filled_cell(v)]
        if len(filled) < MIN_FILLED_CELLS_FOR_HEADER:
            continue
        
        # Calcula score baseado em preenchimento, unicidade e penaliza numéricos
        filled_str = [str(v).strip() for v in filled]
        unique_count = len(set(filled_str))
        numeric_ratio = sum(_looks_numeric(x) for x in filled_str) / len(filled_str) if filled_str else 0
        
        score = len(filled) + UNIQUENESS_BONUS * unique_count - NUMERIC_PENALTY_WEIGHT * numeric_ratio
        
        if score > best_score:
            best_score = score
            best_row = i
    
    wb.close()
    return best_row


@dataclass
class ReadResult:
    """
    Resultado da leitura de uma planilha Excel.
    
    Encapsula tanto o DataFrame lido quanto metadados sobre o processo de leitura,
    permitindo continuar a consolidação mesmo se alguns arquivos falharem.
    
    Attributes:
        df: DataFrame com os dados lidos, ou None se houve falha.
        arquivo: Nome do arquivo (sem caminho completo).
        aba: Nome da aba lida, ou None se houve falha.
        header_row_0based: Índice (0-based) da linha de cabeçalho detectada.
        linhas: Número de linhas de dados lidas (excluindo cabeçalho).
        colunas: Número de colunas no DataFrame.
        status: "OK" se sucesso, "FALHA" caso contrário.
        erro: Mensagem de erro detalhada, ou None se status="OK".
    """
    df: Optional[pd.DataFrame]
    arquivo: str
    aba: Optional[str]
    header_row_0based: Optional[int]
    linhas: int
    colunas: int
    status: str
    erro: Optional[str] = None


def _clean_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove colunas vazias/Unnamed, normaliza nomes e remove linhas vazias.
    
    Aplica as seguintes limpezas:
    1. Normaliza nomes de colunas (remove espaços extras, limita tamanho)
    2. Remove colunas sem nome ou com nome "Unnamed"
    3. Garante unicidade de nomes de coluna (adiciona sufixos _2, _3, ...)
    4. Remove linhas completamente vazias
    5. Remove colunas completamente vazias
    
    Args:
        df: DataFrame a ser limpo.
    
    Returns:
        DataFrame limpo e normalizado.
    """

    # normaliza nomes
    cols = [_normalize_column_name(c) for c in df.columns]

    # remove colunas sem nome ou Unnamed
    keep_mask = []
    cleaned_cols = []
    for c in cols:
        if c == "":
            keep_mask.append(False)
            continue
        if _canon_text(c).startswith("unnamed"):
            keep_mask.append(False)
            continue
        keep_mask.append(True)
        cleaned_cols.append(c)

    df = df.loc[:, keep_mask]

    # unicidade de nomes
    cleaned_cols = _make_unique(cleaned_cols)
    df.columns = cleaned_cols

    # remove linhas totalmente vazias
    df = df.dropna(axis=0, how="all")

    # remove colunas totalmente vazias
    df = df.dropna(axis=1, how="all")

    return df


def ler_planilha_robusta(
    file_path: str,
    preferred_sheet: str,
    auto_detect_header: bool = True,
    header_row_0based: Optional[int] = None,
    read_as_text: bool = True,
    adicionar_auditoria: bool = True,
) -> ReadResult:
    """
    Lê uma planilha Excel e retorna DataFrame + metadados.
    
    Esta função não aborta o processo inteiro em caso de erro, retornando
    um ReadResult com status="FALHA" para permitir consolidação parcial.
    
    Args:
        file_path: Caminho completo do arquivo Excel.
        preferred_sheet: Nome da aba a ler (com tolerância a case/espaços).
        auto_detect_header: Se True, detecta automaticamente a linha de cabeçalho.
        header_row_0based: Índice (0-based) manual do cabeçalho. Ignorado se auto_detect_header=True.
        read_as_text: Se True, lê todas as células como texto (preserva zeros à esquerda).
        adicionar_auditoria: Se True, adiciona colunas de rastreabilidade (arquivo, aba, linha).
    
    Returns:
        ReadResult com DataFrame e metadados da leitura. Status "OK" ou "FALHA".
    
    Examples:
        >>> r = ler_planilha_robusta("data.xlsx", "Sheet1")
        >>> r.status
        'OK'
        >>> r.df.shape
        (100, 15)
    """

    arquivo = os.path.basename(file_path)

    try:
        # Valida existência do arquivo
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Arquivo não encontrado: {file_path}")
        
        aba_real = resolve_sheet_name(file_path, preferred_sheet)

        if header_row_0based is None:
            header = detect_header_row(file_path, aba_real) if auto_detect_header else 1
        else:
            header = header_row_0based

        dtype_arg = "string" if read_as_text else None
        try:
            df = pd.read_excel(
                file_path,
                sheet_name=aba_real,
                header=header,
                engine="openpyxl",
                dtype=dtype_arg,
            )
        except TypeError:
            # pandas sem suporte a dtype="string" em read_excel
            df = pd.read_excel(
                file_path,
                sheet_name=aba_real,
                header=header,
                engine="openpyxl",
                dtype=str if read_as_text else None,
            )

        df = _clean_dataframe(df)

        # auditoria / rastreabilidade usando Enum
        if adicionar_auditoria:
            # linha 1-based no Excel: cabeçalho está em header+1, dados começam em header+2
            linha_ini = header + 2
            df.insert(0, AuditColumn.SOURCE_FILE.value, arquivo)
            df.insert(1, AuditColumn.SOURCE_SHEET.value, aba_real)
            df.insert(2, AuditColumn.HEADER_LINE.value, header + 1)  # 1-based
            df.insert(3, AuditColumn.ORIGINAL_ROW.value, pd.Series(range(linha_ini, linha_ini + len(df)), dtype="Int64"))

        return ReadResult(
            df=df,
            arquivo=arquivo,
            aba=aba_real,
            header_row_0based=header,
            linhas=int(len(df)),
            colunas=int(df.shape[1]),
            status="OK",
        )

    except (FileNotFoundError, PermissionError) as e:
        return ReadResult(
            df=None,
            arquivo=arquivo,
            aba=None,
            header_row_0based=None,
            linhas=0,
            colunas=0,
            status="FALHA",
            erro=f"Erro de acesso ao arquivo: {e}",
        )
    except ValueError as e:
        return ReadResult(
            df=None,
            arquivo=arquivo,
            aba=None,
            header_row_0based=None,
            linhas=0,
            colunas=0,
            status="FALHA",
            erro=f"Erro de validação: {e}",
        )
    except Exception as e:
        return ReadResult(
            df=None,
            arquivo=arquivo,
            aba=None,
            header_row_0based=None,
            linhas=0,
            colunas=0,
            status="FALHA",
            erro=f"Erro inesperado: {e}",
        )


# -----------------------------
# Consolidação
# -----------------------------

def _colunas_final(planilhas: List[pd.DataFrame]) -> List[str]:
    """
    Define colunas finais preservando ordem do primeiro arquivo.
    
    Garante que:
    1. Colunas do primeiro arquivo aparecem primeiro, na ordem original
    2. Colunas adicionais de outros arquivos são adicionadas ao final
    3. Cada coluna aparece apenas uma vez (unicidade preservada)
    
    Args:
        planilhas: Lista de DataFrames a consolidar.
    
    Returns:
        Lista ordenada de nomes de colunas únicos.
    
    Examples:
        >>> df1 = pd.DataFrame({"A": [1], "B": [2]})
        >>> df2 = pd.DataFrame({"B": [3], "C": [4]})
        >>> _colunas_final([df1, df2])
        ['A', 'B', 'C']
    """
    if not planilhas:
        return []

    final: List[str] = []
    seen: set[str] = set()

    def add(c: str) -> None:
        if c not in seen:
            seen.add(c)
            final.append(c)

    for c in planilhas[0].columns:
        add(c)

    for df in planilhas[1:]:
        for c in df.columns:
            add(c)

    return final


def consolidar_planilhas(planilhas: List[pd.DataFrame]) -> pd.DataFrame:
    """
    Consolida uma lista de DataFrames unificando colunas.
    
    Todos os DataFrames são alinhados para ter o mesmo conjunto de colunas,
    com valores NaN onde colunas não existiam no DataFrame original.
    
    Args:
        planilhas: Lista de DataFrames a consolidar.
    
    Returns:
        DataFrame consolidado com todas as linhas de todos os DataFrames,
        com índice resetado (ignore_index=True).
    
    Examples:
        >>> df1 = pd.DataFrame({"A": [1, 2]})
        >>> df2 = pd.DataFrame({"A": [3], "B": [4]})
        >>> consolidar_planilhas([df1, df2])
           A    B
        0  1  NaN
        1  2  NaN
        2  3  4.0
    """
    if not planilhas:
        return pd.DataFrame()

    cols = _colunas_final(planilhas)
    aligned = [df.reindex(columns=cols) for df in planilhas]
    # Filtra DataFrames vazios/all-NA para evitar FutureWarning no concat
    aligned = [df for df in aligned if not df.empty and not df.isna().all(axis=None)]
    if not aligned:
        return pd.DataFrame(columns=cols)
    
    with warnings.catch_warnings():
        warnings.simplefilter("ignore", category=FutureWarning)
        return pd.concat(aligned, ignore_index=True)


# -----------------------------
# Saída Excel (com formatação)
# -----------------------------

def _aplicar_formatacao_worksheet(ws: Worksheet, table_name: str) -> None:
    """
    Aplica formatação profissional ao worksheet Excel.
    
    Aplica as seguintes formatações:
    1. Congela a primeira linha (cabeçalho)
    2. Adiciona auto-filtro em todas as colunas
    3. Ajusta largura das colunas baseado no conteúdo (amostra)
    4. Formata como tabela Excel com estilo profissional
    
    Args:
        ws: Worksheet do openpyxl a formatar.
        table_name: Nome para a tabela Excel (deve ser único no workbook).
    
    Note:
        A amostragem de largura é limitada a WIDTH_SAMPLE_ROWS linhas
        para melhor performance em arquivos grandes.
    """

    if ws.max_row < 1 or ws.max_column < 1:
        return

    # congelar cabeçalho
    ws.freeze_panes = "A2"

    # filtro
    ws.auto_filter.ref = ws.dimensions

    # ajustar larguras (amostra: cabeçalho + primeiras WIDTH_SAMPLE_ROWS linhas)
    max_row_sample = min(ws.max_row, WIDTH_SAMPLE_ROWS + 1)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        for row_idx in range(1, max_row_sample + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        width = min(max_len + COLUMN_PADDING, MAX_COLUMN_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(MIN_COLUMN_WIDTH, width)

    # tabela (se houver ao menos cabeçalho + 1 linha)
    if ws.max_row >= 2:
        last_col = get_column_letter(ws.max_column)
        ref = f"A1:{last_col}{ws.max_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        table.tableStyleInfo = style

        # evita duplicidade se re-salvar
        if not any(t.displayName == table_name for t in ws._tables):
            ws.add_table(table)


def salvar_excel(
    df_dados: pd.DataFrame,
    df_resumo: pd.DataFrame,
    output_path: str,
    sheet_name: str = DEFAULT_SHEET_NAME,
    formatar: bool = True,
) -> None:
    """
    Salva DataFrames consolidados em arquivo Excel formatado.
    
    Cria arquivo Excel com duas abas:
    1. Aba de dados consolidados (com nome configurável)
    2. Aba "Resumo" com metadados da consolidação
    
    Args:
        df_dados: DataFrame com dados consolidados.
        df_resumo: DataFrame com resumo da consolidação (arquivos processados, status, erros).
        output_path: Caminho completo do arquivo de saída (.xlsx).
        sheet_name: Nome da aba de dados (padrão: DEFAULT_SHEET_NAME).
        formatar: Se True, aplica formatação profissional (tabelas, filtros, larguras).
    
    Note:
        O diretório de saída é criado automaticamente se não existir.
    """
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df_dados.to_excel(writer, sheet_name=sheet_name, index=False)
        df_resumo.to_excel(writer, sheet_name="Resumo", index=False)

        if formatar:
            wb = writer.book
            ws_dados = wb[sheet_name]
            ws_resumo = wb["Resumo"]
            _aplicar_formatacao_worksheet(ws_dados, table_name="DadosConsolidados")
            _aplicar_formatacao_worksheet(ws_resumo, table_name="ResumoConsolidacao")


# -----------------------------
# Função principal reutilizável
# -----------------------------

def consolidar(
    arquivos: Iterable[str],
    output_path: str,
    preferred_sheet: str = DEFAULT_SHEET_NAME,
    auto_detect_header: bool = True,
    header_row_0based: Optional[int] = None,
    read_as_text: bool = True,
    adicionar_auditoria: bool = True,
    formatar_saida: bool = True,
    logger: Optional[logging.Logger] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Consolida múltiplos arquivos Excel e salva o resultado.
    
    Função principal que orquestra todo o processo de consolidação:
    1. Lê cada arquivo individualmente (não aborta em caso de erro)
    2. Consolida todos os DataFrames bem-sucedidos
    3. Gera resumo com status de cada arquivo
    4. Salva resultado em Excel formatado
    
    Args:
        arquivos: Iteração de caminhos de arquivos Excel a consolidar.
        output_path: Caminho do arquivo de saída (.xlsx).
        preferred_sheet: Nome da aba a ler (tolerante a case/espaços).
        auto_detect_header: Se True, detecta automaticamente linha de cabeçalho.
        header_row_0based: Índice (0-based) manual do cabeçalho. Ignorado se auto_detect_header=True.
        read_as_text: Se True, lê células como texto (preserva zeros à esquerda).
        adicionar_auditoria: Se True, adiciona colunas de rastreabilidade.
        formatar_saida: Se True, aplica formatação profissional ao Excel.
        logger: Logger customizado. Se None, cria um padrão.
    
    Returns:
        Tupla (df_consolidado, df_resumo) com dados e metadados.
    
    Raises:
        RuntimeError: Se nenhum arquivo foi consolidado com sucesso.
    
    Examples:
        >>> df, resumo = consolidar(
        ...     ["jan.xlsx", "fev.xlsx"],
        ...     "consolidado.xlsx"
        ... )
        >>> df.shape
        (1000, 20)
    """

    log = logger or logging.getLogger("consolidacao")

    resultados: List[ReadResult] = []
    dfs_ok: List[pd.DataFrame] = []

    for p in arquivos:
        r = ler_planilha_robusta(
            file_path=p,
            preferred_sheet=preferred_sheet,
            auto_detect_header=auto_detect_header,
            header_row_0based=header_row_0based,
            read_as_text=read_as_text,
            adicionar_auditoria=adicionar_auditoria,
        )
        resultados.append(r)
        if r.status == "OK" and r.df is not None:
            dfs_ok.append(r.df)
            log.info("OK: %s (%s linhas)", r.arquivo, r.linhas)
        else:
            log.error("FALHA: %s -> %s", r.arquivo, r.erro)

    df_consolidado = consolidar_planilhas(dfs_ok)

    df_resumo = pd.DataFrame(
        [
            {
                "arquivo": r.arquivo,
                "status": r.status,
                "aba": r.aba,
                "header_linha": (r.header_row_0based + 1) if r.header_row_0based is not None else None,
                "linhas": r.linhas,
                "colunas": r.colunas,
                "erro": r.erro,
            }
            for r in resultados
        ]
    )

    if df_consolidado.empty and not any(r.status == "OK" for r in resultados):
        raise RuntimeError(
            "Nenhum arquivo foi consolidado com sucesso.\n"
            "Verifique o LOG para detalhes dos erros e a aba 'Resumo' no arquivo de saída."
        )

    salvar_excel(
        df_dados=df_consolidado,
        df_resumo=df_resumo,
        output_path=output_path,
        sheet_name=preferred_sheet,
        formatar=formatar_saida,
    )

    return df_consolidado, df_resumo


# -----------------------------
# Interface gráfica (Tkinter)
# -----------------------------

AppBase = tk.Tk if tk else object

class App(AppBase):
    """
    Interface gráfica para consolidação de planilhas Excel.
    
    Aplicação Tkinter que permite ao usuário:
    - Selecionar múltiplos arquivos Excel
    - Configurar opções de leitura e consolidação
    - Processar arquivos com barra de progresso
    - Salvar resultado consolidado formatado
    """
    
    def __init__(self) -> None:
        """Inicializa a aplicação e constrói interface."""
        super().__init__()

        self.title("Consolidar Relatório Base")
        self.geometry("780x520")
        self.minsize(780, 520)

        self.selected_files: List[str] = []

        # Opções
        self.var_sheet = tk.StringVar(value=DEFAULT_SHEET_NAME)
        self.var_auto_header = tk.BooleanVar(value=True)
        self.var_header_manual = tk.IntVar(value=2)  # 1-based para o usuário
        self.var_read_text = tk.BooleanVar(value=True)
        self.var_auditoria = tk.BooleanVar(value=True)
        self.var_formatar = tk.BooleanVar(value=True)

        self._build_ui()

    def _build_ui(self) -> None:
        """Constrói todos os componentes da interface gráfica."""
        frm_top = ttk.Frame(self, padding=10)
        frm_top.pack(fill="x")

        # Linha: aba
        ttk.Label(frm_top, text="Nome da aba:").grid(row=0, column=0, sticky="w")
        ttk.Entry(frm_top, textvariable=self.var_sheet, width=30).grid(row=0, column=1, sticky="w", padx=6)

        # Linha: header
        ttk.Checkbutton(
            frm_top,
            text="Detectar cabeçalho automaticamente",
            variable=self.var_auto_header,
            command=self._toggle_header_manual,
        ).grid(row=0, column=2, sticky="w", padx=10)

        ttk.Label(frm_top, text="Cabeçalho (linha):").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.spn_header = ttk.Spinbox(frm_top, from_=1, to=200, textvariable=self.var_header_manual, width=6)
        self.spn_header.grid(row=1, column=1, sticky="w", padx=6, pady=(8, 0))

        # Opções
        frm_opts = ttk.Frame(self, padding=(10, 0, 10, 10))
        frm_opts.pack(fill="x")

        ttk.Checkbutton(frm_opts, text="Ler tudo como texto (preserva zeros à esquerda)", variable=self.var_read_text).grid(
            row=0, column=0, sticky="w"
        )
        ttk.Checkbutton(frm_opts, text="Adicionar colunas de auditoria (origem/linha)", variable=self.var_auditoria).grid(
            row=1, column=0, sticky="w", pady=(4, 0)
        )
        ttk.Checkbutton(frm_opts, text="Formatar saída (filtro, congelar, larguras, tabela)", variable=self.var_formatar).grid(
            row=2, column=0, sticky="w", pady=(4, 0)
        )

        # Botões
        frm_btn = ttk.Frame(self, padding=10)
        frm_btn.pack(fill="x")

        self.btn_select = ttk.Button(frm_btn, text="Selecionar arquivos (.xlsx)", command=self._select_files)
        self.btn_select.pack(side="left")

        self.btn_run = ttk.Button(frm_btn, text="Consolidar e salvar", command=self._run)
        self.btn_run.pack(side="left", padx=8)

        # Lista de arquivos
        frm_list = ttk.Frame(self, padding=(10, 0, 10, 10))
        frm_list.pack(fill="both", expand=True)

        ttk.Label(frm_list, text="Arquivos selecionados:").pack(anchor="w")

        self.listbox = tk.Listbox(frm_list, height=10)
        self.listbox.pack(fill="both", expand=True, pady=(4, 0))

        # Barra de progresso
        frm_status = ttk.Frame(self, padding=10)
        frm_status.pack(fill="x")

        self.progress = ttk.Progressbar(frm_status, mode="determinate")
        self.progress.pack(fill="x")

        self.lbl_status = ttk.Label(frm_status, text="")
        self.lbl_status.pack(anchor="w", pady=(6, 0))

        self._toggle_header_manual()

    def _toggle_header_manual(self) -> None:
        """Habilita/desabilita spinbox de cabeçalho manual baseado em auto-detect."""
        # Se auto-detect está ligado, desabilita spinbox
        state = "disabled" if self.var_auto_header.get() else "normal"
        self.spn_header.configure(state=state)

    def _select_files(self) -> None:
        """Abre diálogo para seleção de arquivos Excel."""
        paths = filedialog.askopenfilenames(
            title="Selecione as planilhas mensais a consolidar",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if not paths:
            return

        self.selected_files = list(paths)
        self.listbox.delete(0, tk.END)
        for p in self.selected_files:
            self.listbox.insert(tk.END, p)

        self.lbl_status.configure(text=f"{len(self.selected_files)} arquivo(s) selecionado(s).")


    def _prepare_consolidation_args(self, output_path: str, logger: logging.Logger) -> Dict[str, Any]:
        """
        Prepara argumentos de configuração para consolidação.
        
        Args:
            output_path: Caminho do arquivo de saída.
            logger: Logger configurado para o processo.
        
        Returns:
            Dicionário com todos os argumentos necessários para consolidação.
        """
        return {
            "arquivos": self.selected_files,
            "output_path": output_path,
            "preferred_sheet": self.var_sheet.get().strip() or DEFAULT_SHEET_NAME,
            "auto_detect_header": self.var_auto_header.get(),
            "header_row_0based": None if self.var_auto_header.get() else max(0, int(self.var_header_manual.get()) - 1),
            "read_as_text": self.var_read_text.get(),
            "adicionar_auditoria": self.var_auditoria.get(),
            "formatar_saida": self.var_formatar.get(),
            "logger": logger,
        }
    
    def _setup_logger(self, output_path: str) -> Tuple[logging.Logger, str]:
        """
        Configura logger para arquivo de log.
        
        Args:
            output_path: Caminho do arquivo de saída (.xlsx).
        
        Returns:
            Tupla (logger, log_path) com logger configurado e caminho do log.
        """
        log_path = os.path.splitext(output_path)[0] + ".log"
        logger = logging.getLogger("consolidacao")
        logger.setLevel(logging.INFO)
        logger.handlers.clear()
        fh = logging.FileHandler(log_path, encoding="utf-8")
        fh.setFormatter(logging.Formatter("%(asctime)s | %(levelname)s | %(message)s"))
        logger.addHandler(fh)
        return logger, log_path
    
    def _configure_ui_for_processing(self) -> None:
        """Desabilita botões e configura UI para modo de processamento."""
        self.btn_select.configure(state="disabled")
        self.btn_run.configure(state="disabled")
        self.progress.configure(maximum=len(self.selected_files), value=0)
        self.lbl_status.configure(text="Iniciando...")
    
    def _process_files_with_progress(
        self, 
        args: Dict[str, Any]
    ) -> Tuple[List[ReadResult], List[pd.DataFrame]]:
        """
        Processa arquivos com atualização de progresso na UI.
        
        Args:
            args: Dicionário com argumentos de consolidação.
        
        Returns:
            Tupla (results, dfs_ok) com resultados de todos os arquivos e
            lista de DataFrames processados com sucesso.
        """
        results: List[ReadResult] = []
        dfs_ok: List[pd.DataFrame] = []
        
        for idx, p in enumerate(self.selected_files, start=1):
            r = ler_planilha_robusta(
                file_path=p,
                preferred_sheet=args["preferred_sheet"],
                auto_detect_header=args["auto_detect_header"],
                header_row_0based=args["header_row_0based"],
                read_as_text=args["read_as_text"],
                adicionar_auditoria=args["adicionar_auditoria"],
            )
            results.append(r)
            if r.status == "OK" and r.df is not None:
                dfs_ok.append(r.df)
            args["logger"].info("%s: %s", r.status, r.arquivo)
            
            # Atualiza UI na thread principal
            self.after(0, self._set_progress, idx, f"Processado: {r.arquivo} ({r.status})")
        
        return results, dfs_ok
    
    def _generate_summary_dataframe(self, results: List[ReadResult]) -> pd.DataFrame:
        """
        Gera DataFrame de resumo a partir dos resultados.
        
        Args:
            results: Lista de ReadResult de todos os arquivos processados.
        
        Returns:
            DataFrame com resumo da consolidação.
        """
        return pd.DataFrame(
            [
                {
                    "arquivo": r.arquivo,
                    "status": r.status,
                    "aba": r.aba,
                    "header_linha": (r.header_row_0based + 1) if r.header_row_0based is not None else None,
                    "linhas": r.linhas,
                    "colunas": r.colunas,
                    "erro": r.erro,
                }
                for r in results
            ]
        )
    
    def _create_success_message(
        self, 
        results: List[ReadResult], 
        df_consolidado: pd.DataFrame, 
        output_path: str, 
        log_path: str
    ) -> str:
        """
        Cria mensagem de sucesso com estatísticas.
        
        Args:
            results: Lista de ReadResult.
            df_consolidado: DataFrame consolidado.
            output_path: Caminho do arquivo de saída.
            log_path: Caminho do arquivo de log.
        
        Returns:
            Mensagem formatada para exibição.
        """
        ok_count = sum(r.status == "OK" for r in results)
        fail_count = sum(r.status != "OK" for r in results)
        total_rows = int(df_consolidado.shape[0])
        
        return (
            f"Consolidação concluída.\n\n"
            f"Arquivos OK: {ok_count}\n"
            f"Arquivos com falha: {fail_count}\n"
            f"Linhas consolidadas: {total_rows}\n\n"
            f"Saída: {output_path}\n"
            f"Log: {log_path}"
        )

    def _run(self) -> None:
        """
        Executa consolidação em thread separada com feedback de progresso.
        
        Valida entrada, configura UI, cria worker thread para processar
        arquivos sem bloquear interface, e exibe resultado ao usuário.
        """
        if not self.selected_files:
            messagebox.showwarning("Atenção", "Selecione ao menos um arquivo .xlsx.")
            return

        output_path = filedialog.asksaveasfilename(
            title="Salvar arquivo consolidado como",
            defaultextension=".xlsx",
            filetypes=[("Arquivos Excel", "*.xlsx")],
        )
        if not output_path:
            return

        logger, log_path = self._setup_logger(output_path)
        self._configure_ui_for_processing()
        args = self._prepare_consolidation_args(output_path, logger)

        def worker() -> None:
            """Worker thread para processamento de arquivos."""
            try:
                # Processa arquivos com progresso
                results, dfs_ok = self._process_files_with_progress(args)
                
                # Consolida e gera resumo
                df_consolidado = consolidar_planilhas(dfs_ok)
                df_resumo = self._generate_summary_dataframe(results)

                # Valida se houve sucesso
                if df_consolidado.empty and not any(r.status == "OK" for r in results):
                    raise RuntimeError(
                        "Nenhum arquivo foi consolidado com sucesso.\n"
                        "Verifique o LOG e a aba 'Resumo'."
                    )

                # Salva resultado
                salvar_excel(
                    df_dados=df_consolidado,
                    df_resumo=df_resumo,
                    output_path=args["output_path"],
                    sheet_name=args["preferred_sheet"],
                    formatar=args["formatar_saida"],
                )

                # Mensagem de sucesso
                msg = self._create_success_message(results, df_consolidado, args["output_path"], log_path)
                self.after(0, self._finish_ok, msg)

            except (FileNotFoundError, PermissionError) as e:
                self.after(0, self._finish_error, f"Erro de acesso ao arquivo: {e}", log_path)
            except ValueError as e:
                self.after(0, self._finish_error, f"Erro de validação: {e}", log_path)
            except Exception as e:
                self.after(0, self._finish_error, f"Erro inesperado: {e}", log_path)

        threading.Thread(target=worker, daemon=True).start()

    def _set_progress(self, value: int, status: str) -> None:
        """Atualiza barra de progresso e mensagem de status.
        
        Args:
            value: Valor atual do progresso (número de arquivos processados).
            status: Mensagem de status a exibir.
        """
        self.progress.configure(value=value)
        self.lbl_status.configure(text=status)

    def _finish_ok(self, msg: str) -> None:
        """Finaliza processamento com sucesso e exibe mensagem.
        
        Args:
            msg: Mensagem de sucesso com estatísticas da consolidação.
        """
        self.btn_select.configure(state="normal")
        self.btn_run.configure(state="normal")
        self.lbl_status.configure(text="Concluído.")
        messagebox.showinfo("Consolidação concluída", msg)

    def _finish_error(self, err: str, log_path: str) -> None:
        """Finaliza processamento com erro e exibe mensagem.
        
        Args:
            err: Mensagem de erro descritiva.
            log_path: Caminho do arquivo de log para referência.
        """
        self.btn_select.configure(state="normal")
        self.btn_run.configure(state="normal")
        self.lbl_status.configure(text="Erro.")
        messagebox.showerror(
            "Erro na consolidação",
            f"Ocorreu um erro:\n{err}\n\nVerifique o log em:\n{log_path}",
        )


def main() -> None:
    """Ponto de entrada da aplicação GUI."""
    if tk is None:
        print("Tkinter não disponível. Execute em um ambiente com suporte a GUI ou use a interface web.")
        return
    app = App()
    app.mainloop()


if __name__ == "__main__":
    main()
